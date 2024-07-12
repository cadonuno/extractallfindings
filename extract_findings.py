# Credit to Ricardo P for this script

import sys
import requests
import argparse
from openpyxl import Workbook 
import time
import os.path
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
from veracode_api_signing.credentials import get_credentials
import xml.dom.minidom as xml
import base64

json_header = {
    "User-Agent": "Findings extractor",
    "Content-Type": "application/json"
}


severity_map = { 
    0: "Informational",
    1: "Very Low",
    2: "Low",
    3: "Medium",
    4: "High",
    5: "Very High",
}

def print_help():
    print("""extractfindings.py -t <target_file(.xlsx)> -a <application_guid> [-s] [-d]
        Gets ALL findings for a specific application profile, including latest scan name and date, and saves it to <target_file>""")
    print("Optional arguments: ")
    print(" -d: set to enable fetching of DAST results")
    print(" -s: set to enable fetching of SCA results")
    print(" -v: to output verbose logs")
    sys.exit()

def get_rest_api_base():
    api_key_id, api_key_secret = get_credentials()
    if api_key_id.startswith("vera01"):
        return "https://api.veracode.eu/"
    else:
        return "https://api.veracode.com/"
    
def get_xml_api_base():
    api_key_id, api_key_secret = get_credentials()
    if api_key_id.startswith("vera01"):
        return "https://analysiscenter.veracode.eu/api/5.0/"
    else:
        return "https://analysiscenter.veracode.com/api/5.0/"

def handle_throttling():
    print("429 returned, waiting 1 minute")
    time.sleep(60)

def has_more_pages(body):
    return body["page"]["number"]+1 < body["page"]["total_pages"]

def get_findings_for_app_and_scan_type(application_guid, page, rest_api_base, scan_type, verbose):
    print(f"Getting findings for application {application_guid} - page {page}")
    path = f"{rest_api_base}appsec/v2/applications/{application_guid}/findings?scan_type={scan_type}&page={page}"

    if verbose:
        print(f"Calling API at {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_header)

    body = response.json()
    if verbose:
        print(f"status code {response.status_code}")
        if body:
            print(body)
    if response.status_code == 200:
        print(f"Successfully obtained {scan_type} findings page {page}")
        if "_embedded" in body and "findings" in body["_embedded"]:
            findings = body["_embedded"]["findings"]
            if has_more_pages(body):
                return findings + get_findings_for_app_and_scan_type(application_guid, page+1, rest_api_base, scan_type, verbose)
            else:
                return findings
        return []
    elif response.status_code == 429:
        handle_throttling()
        return get_findings_for_app_and_scan_type(application_guid, page, rest_api_base, scan_type, verbose)
    else:
        print(f"Unable to obtain {scan_type} findings: {response.status_code}")
        return []

def get_application_results(application_guid, rest_api_base, is_sast, is_dast, is_sca, verbose):
    if is_sast:
        sast_findings = get_findings_for_app_and_scan_type(application_guid, 0, rest_api_base, "STATIC", verbose)
    else:
        sast_findings = []
    if is_dast:
        dast_findings = get_findings_for_app_and_scan_type(application_guid, 0, rest_api_base, "DYNAMIC", verbose)
    else:
        dast_findings = []
    if is_sca: 
        sca_findings = get_findings_for_app_and_scan_type(application_guid, 0, rest_api_base, "SCA", verbose)
    else:
        sca_findings = []

    return {
        "sast": sast_findings, 
        "dast": dast_findings, 
        "sca": sca_findings
        }

def get_exploitability(cve_node):
    if not 'exploitability' in cve_node or not cve_node['exploitability'] or not "epss_score" in cve_node['exploitability']:
        return { 
            "score": "Unavailable",
            "percentile": "Unavailable"
        }
    return { 
            "score": cve_node['exploitability']['epss_score'],  
            "percentile": cve_node['exploitability']['epss_percentile']
        }

def try_decode(element):
    try:
        return base64.b64decode(element)
    except Exception:
        return element

def write_row(worksheet, row, content):
    column = 1
    for value in content:
        worksheet.cell(row=row, column=column).value = value
        column+=1

def save_sast(sast_findings, worksheet):
    write_row(worksheet, 1, ["Issue ID", "Description", "Violates Policy?", "First Found Date", "Status", "Resolution", "Severity", "CWE", "Finding Category", "Module Name", "File Path", "Line"])
    row = 2
    for finding in sast_findings:
        write_row(worksheet, row, [
                finding['issue_id'], try_decode(finding['description']), finding['violates_policy'], finding['finding_status']['first_found_date'], finding['finding_status']['status'], 
                finding['finding_status']['resolution'], severity_map[finding['finding_details']['severity']], 
                "CWE " + str(finding['finding_details']['cwe']['id']) + ": " + finding['finding_details']['cwe']['name'], finding['finding_details']['finding_category']['name'], 
                finding['finding_details']['module'], finding['finding_details']['file_path'], finding['finding_details']['file_line_number'] 
            ])
        row+=1

def save_dast(dast_findings, worksheet):
    write_row(worksheet, 1, ["Issue ID", "Description", "Violates Policy?", "First Found Date", "Status", "Resolution", "Severity", "CWE", "Finding Category", "URL", "Attack Vector", "Vulnerable Parameter"])
    row = 2
    for finding in dast_findings:
        write_row(worksheet, row, [
                finding['issue_id'], try_decode(finding['description']), finding['violates_policy'], finding['finding_status']['first_found_date'], finding['finding_status']['status'], 
                finding['finding_status']['resolution'], severity_map[finding['finding_details']['severity']], 
                "CWE " + str(finding['finding_details']['cwe']['id']) + ": " + finding['finding_details']['cwe']['name'], finding['finding_details']['finding_category']['name'], 
                finding['finding_details']['url'], finding['finding_details']['attack_vector'], finding['finding_details']['vulnerable_parameter'] if 'vulnerable_parameter' in finding['finding_details'] else ''
            ])
        row+=1

def save_sca(sca_findings, worksheet):
    write_row(worksheet, 1, ["Component File", "Component Version", "Description", "Violates Policy?", "First Found Date", "Status", "Resolution", "Severity", "EPSS Score", "EPSS Percentile", "CVE", "CVSS2", "CVSS3"])
    row = 2
    for finding in sca_findings:
        exploitability = get_exploitability(finding['finding_details']['cve'])
        write_row(worksheet, row, [finding['finding_details']['component_filename'], finding['finding_details']['version'],
            try_decode(finding['description']), finding['violates_policy'], finding['finding_status']['first_found_date'], finding['finding_status']['status'], 
            finding['finding_status']['resolution'], severity_map[finding['finding_details']['severity']], exploitability["score"], exploitability["percentile"],
            finding['finding_details']['cve']['name'], finding['finding_details']['cve']['cvss'], 
            finding['finding_details']['cve']['cvss3']['score']])
        row+=1

def save_to_excel(findings, file_name):
    directory = os.path.dirname(file_name)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)
    if findings:
            workbook = Workbook()
            workbook.remove(workbook.active)
            if findings['sast']:
                save_sast(findings['sast'], workbook.create_sheet(title="SAST"))
            if findings['dast']:
                save_dast(findings['dast'], workbook.create_sheet(title="DAST"))
            if findings['sca']:
                save_sca(findings['sca'], workbook.create_sheet(title="SCA"))

            workbook.save(file_name)
    else:
        print(f"ERROR: No findings found")

def save_scan_results(rest_api_base, application_guid, target_file, is_sast, is_dast, is_sca, verbose):
    _, extension = os.path.splitext(target_file)
    if not extension or extension.lower() != ".xlsx":
        print(f"ERROR: File name '{target_file}' needs to be a XLSX file.")
        sys.exit(-1)

    save_to_excel(get_application_results(application_guid, rest_api_base, is_sast, is_dast, is_sca, verbose), target_file)

def main():
    try:        

        parser = argparse.ArgumentParser(
        description='This script will create an excel file with a summary of all your SAST scans')
        
        parser.add_argument('-t', '--target', help='XLSX file to save results')
        parser.add_argument('-s', '--sast', action='store_true', help='Set to enable fetching of SAST results')
        parser.add_argument('-d', '--dast', action='store_true', help='Set to enable fetching of DAST results')
        parser.add_argument('-c', '--sca', action='store_true', help='Set to enable fetching of SCA results')
        parser.add_argument('-a', '--application_guid', help='Set to enable fetching of DAST results')
        parser.add_argument('-v', '--verbose', action='store_true', help='Set to enable verbose logging')

        args = parser.parse_args()

        target_file = args.target
        is_sast = args.sast
        is_dast = args.dast
        is_sca = args.sca
        
        application_guid = args.application_guid
        verbose = args.verbose

        rest_api_base = get_rest_api_base()
        save_scan_results(rest_api_base, application_guid, target_file, is_sast, is_dast, is_sca, verbose)

    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)


if __name__ == "__main__":
    main()
